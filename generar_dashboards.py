""" =============================================================
 GENERADOR AUTOMÁTICO DE DASHBOARDS - KAIBIL / SEGURA
 Hik-Central Professional → GitHub Pages
 Configurado para: Kaibiluy / d_tel
 v2.2 — autenticación SHA-256 por dashboard
============================================================= """
import os, re, sys, subprocess
from datetime import datetime
from collections import defaultdict

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
CARPETA_EXCEL  = r"C:\Users\d_tel\Documents\Hik-Central"
CARPETA_REPO   = r"C:\Users\d_tel\Documents\dashboards"
GITHUB_USUARIO = "Kaibiluy"
MAX_DIAS       = 7

# ─────────────────────────────────────────────
# AUTENTICACIÓN SHA-256
# ─────────────────────────────────────────────
HASH_KAIBIL = "4f1f81565875bfb546ce204acb7953474faf941d36076fd4f9eeef4d08784de0"
HASH_SEGURA  = "4953cc2b28a1d908c1a19fd082254c78e60eb3e3e195aa7bfd951c0bdb1068b0"

# filename → (hash, session_key, logo, color, sublabel)
DASHBOARD_AUTH = {
    'index.html':          (HASH_KAIBIL, 'auth_kaibil', 'KAIBIL', '#4a9eff', 'Seguridad Privada'),
    'segura.html':         (HASH_SEGURA,  'auth_segura',  'SEGURA', '#a855f7', 'Monitoreo y Vigilancia'),
    'punta-del-este.html': (HASH_SEGURA,  'auth_pde',     'SEGURA', '#4a9eff', 'Punta del Este'),
}

AUTH_CSS = """
/* ── Auth overlay ── */
#auth-overlay{position:fixed;inset:0;background:#0f1117;z-index:9999;
  display:flex;align-items:center;justify-content:center}
.auth-box{background:#1a1f2e;border:1px solid #2d3550;border-radius:16px;
  padding:44px 48px;display:flex;flex-direction:column;align-items:center;gap:20px;
  box-shadow:0 8px 40px #00000088;min-width:320px}
.auth-logo{font-size:2rem;font-weight:900;letter-spacing:3px}
.auth-sub{font-size:0.82rem;color:#556;letter-spacing:1px;margin-top:-12px}
.auth-label{font-size:0.78rem;color:#8899bb;align-self:flex-start}
.auth-input{width:100%;background:#0f1117;border:1px solid #2d3550;border-radius:8px;
  padding:10px 14px;color:#e0e0e0;font-size:0.95rem;outline:none;transition:border .2s}
.auth-input:focus{border-color:#4a9eff}
.auth-btn{width:100%;background:#1a3a6e;color:#fff;border:none;border-radius:8px;
  padding:11px;font-size:0.9rem;font-weight:700;cursor:pointer;letter-spacing:1px;
  transition:background .2s}
.auth-btn:hover{background:#2a5aae}
.auth-err{color:#ef4444;font-size:0.78rem;display:none;margin-top:-8px}
"""

def auth_block(hash_val, session_key, logo_txt, logo_color, sub_label):
    return f"""
<div id="auth-overlay">
  <div class="auth-box">
    <div class="auth-logo" style="color:{logo_color}">{logo_txt}</div>
    <div class="auth-sub">{sub_label}</div>
    <div class="auth-label">Contraseña de acceso</div>
    <input id="auth-pwd" class="auth-input" type="password"
           placeholder="••••••••••" onkeydown="if(event.key==='Enter')checkAuth()">
    <button class="auth-btn" onclick="checkAuth()">INGRESAR</button>
    <span class="auth-err" id="auth-err">Contraseña incorrecta</span>
  </div>
</div>
<script>
(function(){{
  var SK='{session_key}';
  var HASH='{hash_val}';
  if(sessionStorage.getItem(SK)==='1'){{
    document.getElementById('auth-overlay').style.display='none';return;
  }}
  async function sha256(s){{
    var b=await crypto.subtle.digest('SHA-256',new TextEncoder().encode(s));
    return Array.from(new Uint8Array(b)).map(x=>x.toString(16).padStart(2,'0')).join('');
  }}
  window.checkAuth=async function(){{
    var h=await sha256(document.getElementById('auth-pwd').value);
    if(h===HASH){{
      sessionStorage.setItem(SK,'1');
      document.getElementById('auth-overlay').style.display='none';
    }}else{{
      var e=document.getElementById('auth-err');
      e.style.display='block';
      document.getElementById('auth-pwd').value='';
      document.getElementById('auth-pwd').style.borderColor='#ef4444';
      setTimeout(function(){{e.style.display='none';
        document.getElementById('auth-pwd').style.borderColor='#2d3550';}},2500);
    }}
  }};
}})();
</script>"""

# ─────────────────────────────────────────────
# CLASIFICACIÓN
# ─────────────────────────────────────────────
KC_EXACT = {
    'Constancio Vigil', 'Guillemette 2', 'Guani y Delgado',
    'Rostand', 'NVR KAIBIL EMPRESA'
}

def classify(area):
    a = area.strip()
    if a.upper() == 'BAJAS': return None
    if a.startswith('KS'): return 'KS'
    if a.startswith('KC') or a.lower().startswith('grupo') or a in KC_EXACT: return 'KC'
    if a.startswith('MK'): return 'MK'
    if a.startswith('MS'): return 'MS'
    if a.startswith('SG'): return 'SG'
    if a.startswith('SP'): return 'SP'
    if a.startswith('SR'): return 'SR'
    return 'Otros'

# ─────────────────────────────────────────────
# DETECTAR TIPO DE EXCEL
# ─────────────────────────────────────────────
def es_reporte_camara(ruta):
    return os.path.basename(ruta).lower().startswith('cámara_') or \
           os.path.basename(ruta).lower().startswith('camara_')

def fecha_de_nombre(nombre):
    m = re.search(r'(\d{14})', nombre)
    if m: return datetime.strptime(m.group(1), '%Y%m%d%H%M%S')
    m = re.search(r'(\d{8})', nombre)
    if m: return datetime.strptime(m.group(1), '%Y%m%d')
    return datetime.today()

# ─────────────────────────────────────────────
# GOOGLE DRIVE
# ─────────────────────────────────────────────
GOOGLE_DRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "1fayZgQE5zRmpo9RchY7Ti0nZKU7qeHmS")

def descargar_desde_drive(carpeta_destino, max_archivos=MAX_DIAS):
    try:
        from googleapiclient.discovery import build
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        import pickle as pkl

        SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        token_path  = os.path.join(os.path.dirname(__file__), 'token_drive.pkl')
        creds_path  = os.path.join(os.path.dirname(__file__), 'credentials_drive.json')
        creds = None

        refresh_token  = os.environ.get('GDRIVE_REFRESH_TOKEN')
        client_id      = os.environ.get('GDRIVE_CLIENT_ID')
        client_secret  = os.environ.get('GDRIVE_CLIENT_SECRET')

        if refresh_token and client_id and client_secret:
            from google.oauth2.credentials import Credentials as OAuthCreds
            creds = OAuthCreds(token=None, refresh_token=refresh_token,
                               token_uri='https://oauth2.googleapis.com/token',
                               client_id=client_id, client_secret=client_secret,
                               scopes=SCOPES)
            creds.refresh(Request())
        elif os.path.exists(token_path):
            with open(token_path, 'rb') as tk: creds = pkl.load(tk)
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    if not os.path.exists(creds_path):
                        print(" ⚠ No se encontró credentials_drive.json. Fallback local.")
                        return []
                    flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
                    creds = flow.run_local_server(port=0)
                    with open(token_path, 'wb') as tk: pkl.dump(creds, tk)
        else:
            if not os.path.exists(creds_path):
                print(" ⚠ No se encontró credentials_drive.json. Fallback local.")
                return []
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
            creds = flow.run_local_server(port=0)
            with open(token_path, 'wb') as tk: pkl.dump(creds, tk)

        service = build('drive', 'v3', credentials=creds)
        if not os.path.exists(carpeta_destino): os.makedirs(carpeta_destino)
        rutas = []

        results = service.files().list(
            q=f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and name contains 'Resource' "
              f"and name contains '.xlsx' and trashed=false",
            orderBy='name desc', pageSize=max_archivos, fields='files(id, name)'
        ).execute()
        for archivo in results.get('files', []):
            ruta_local = os.path.join(carpeta_destino, archivo['name'])
            _descargar_archivo(service, archivo, ruta_local)
            rutas.append(ruta_local)

        results2 = service.files().list(
            q=f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and ("
              f"name contains 'Cámara_' or name contains 'Camara_') "
              f"and name contains '.xlsx' and trashed=false",
            orderBy='name desc', pageSize=10, fields='files(id, name)'
        ).execute()
        camara_files = results2.get('files', [])
        if camara_files:
            print(f" 📷 Encontrados {len(camara_files)} reporte(s) manual(es):")
            for archivo in camara_files:
                ruta_local = os.path.join(carpeta_destino, archivo['name'])
                _descargar_archivo(service, archivo, ruta_local)
                rutas.append(ruta_local)

        if not rutas:
            print(" ⚠ No se encontraron archivos en Drive. Fallback local.")
            return []
        return rutas

    except ImportError:
        print(" Instalando librerías Google Drive...")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install',
            'google-api-python-client', 'google-auth-httplib2', 'google-auth-oauthlib'])
        print(" ✓ Instalado. Volvé a ejecutar.")
        sys.exit(0)
    except Exception as e:
        print(f" ⚠ Error Drive: {e}. Fallback local.")
        return []

def _descargar_archivo(service, archivo, ruta_local):
    from googleapiclient.http import MediaIoBaseDownload
    import io
    print(f" Archivo en Drive: {archivo['name']}")
    if not os.path.exists(ruta_local):
        request = service.files().get_media(fileId=archivo['id'])
        fh = io.FileIO(ruta_local, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done: _, done = downloader.next_chunk()
        print(f" ✓ Descargado")
    else:
        print(f" ✓ Ya existe localmente")

# ─────────────────────────────────────────────
# ENCONTRAR EXCELS
# ─────────────────────────────────────────────
def encontrar_excels(carpeta):
    rutas_drive = descargar_desde_drive(carpeta, MAX_DIAS)
    if rutas_drive:
        candidatos = rutas_drive
    else:
        if not os.path.exists(carpeta): os.makedirs(carpeta)
        candidatos = [
            os.path.join(carpeta, f) for f in os.listdir(carpeta)
            if f.endswith('.xlsx') and '__1_' not in f and
               (f.lower().replace(' ','_').startswith('resource_online_and_offline_log') or
                f.lower().startswith('cámara_') or f.lower().startswith('camara_'))
        ]
    if not candidatos:
        print(f"\n ERROR: No se encontró ningún archivo Excel en:\n {carpeta}")
        sys.exit(1)

    logs, camaras = {}, {}
    for ruta in candidatos:
        nombre   = os.path.basename(ruta)
        fecha_dt = fecha_de_nombre(nombre)
        dia_key  = fecha_dt.strftime('%Y%m%d')
        if es_reporte_camara(ruta):
            if dia_key not in camaras or fecha_dt > camaras[dia_key][0]:
                camaras[dia_key] = (fecha_dt, ruta)
        else:
            logs[dia_key] = (fecha_dt, ruta)

    resultado = []
    for dia_key, (fecha_dt, ruta_log) in logs.items():
        if dia_key in camaras:
            _, ruta_cam = camaras[dia_key]
            print(f" 📷 {dia_key[:4]}/{dia_key[4:6]}/{dia_key[6:]} — reporte manual: {os.path.basename(ruta_cam)}")
            resultado.append((fecha_dt, ruta_cam))
        else:
            resultado.append((fecha_dt, ruta_log))
    for dia_key, (fecha_dt, ruta_cam) in camaras.items():
        if dia_key not in logs:
            print(f" 📷 {dia_key[:4]}/{dia_key[4:6]}/{dia_key[6:]} — solo manual: {os.path.basename(ruta_cam)}")
            resultado.append((fecha_dt, ruta_cam))

    resultado.sort(key=lambda x: x[0])
    return resultado[-MAX_DIAS:]

# ─────────────────────────────────────────────
# LEER EXCEL
# ─────────────────────────────────────────────
def leer_excel(ruta):
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
        import openpyxl

    fecha_dt   = fecha_de_nombre(os.path.basename(ruta))
    fecha      = fecha_dt.strftime('%d/%m/%Y')
    area_stats = defaultdict(lambda: {'total': 0, 'offline': 0, 'client': '', 'cameras': []})
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    if es_reporte_camara(ruta):
        print(f" 📷 Leyendo reporte manual: {os.path.basename(ruta)}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 8: continue
            nombre  = str(row[0]).strip() if row[0] else ''
            address = str(row[2]).strip() if row[2] else ''
            area    = str(row[3]).strip() if row[3] else ''
            estado  = str(row[7]).strip() if row[7] else ''
            client  = classify(area)
            if client is None or not area: continue
            area_stats[area]['total']  += 1
            area_stats[area]['client']  = client
            is_offline = 'en línea' not in estado.lower()
            area_stats[area]['cameras'].append({'nombre': nombre, 'address': address, 'offline': is_offline})
            if is_offline: area_stats[area]['offline'] += 1
    else:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= 8: continue
            area    = str(row[2]).strip() if row[2] else ''
            status  = str(row[3]).strip() if row[3] else ''
            nombre  = str(row[0]).strip() if row[0] else ''
            address = str(row[1]).strip() if row[1] else ''
            client  = classify(area)
            if client is None: continue
            area_stats[area]['total']  += 1
            area_stats[area]['client']  = client
            area_stats[area]['cameras'].append({'nombre': nombre, 'address': address, 'offline': status == 'Offline'})
            if status == 'Offline': area_stats[area]['offline'] += 1

    areas = []
    for area, v in area_stats.items():
        pct = round(v['offline'] / v['total'] * 100, 1) if v['total'] else 0
        areas.append({'client': v['client'], 'area': area,
                      'offline': v['offline'], 'total': v['total'],
                      'pct': pct, 'cameras': v['cameras']})
    return areas, fecha

# ─────────────────────────────────────────────
# HTML helpers
# ─────────────────────────────────────────────
def sem(pct):
    if pct == 0:  return ('sem-green',  'pct-green',  '#22c55e')
    if pct < 20:  return ('sem-yellow', 'pct-yellow', '#f59e0b')
    return              ('sem-red',    'pct-red',    '#ef4444')

def area_row(a, show_cameras=False):
    sc, pc, bc = sem(a['pct'])
    pct_str = f"{a['pct']}%" if a['pct'] > 0 else "0%"
    uid = re.sub(r'[^a-z0-9]', '_', a['area'].lower())[:30]
    cam_panel = ''
    cursor = ''
    if show_cameras and a.get('cameras'):
        cams_sorted = sorted(a['cameras'], key=lambda c: (0 if c['offline'] else 1, c['nombre']))
        cam_items = ''.join(
            f'<div class="cam-item"><div class="cam-dot {"cam-off" if c["offline"] else "cam-on"}"></div>'
            f'<span class="cam-name {"cam-name-off" if c["offline"] else ""}">{c["nombre"]}'
            f'<span class="cam-addr"> · {c["address"]}</span></span></div>'
            for c in cams_sorted
        )
        cam_panel = f'<div class="cam-panel" id="cams_{uid}">{cam_items}</div>'
        cursor = 'cursor:pointer;'
    return f"""
 <div class="area-wrap">
  <div class="area-row" style="{cursor}" onclick="{'toggleCams(\'' + uid + '\')' if show_cameras and a.get('cameras') else ''}">
   <div class="semaforo {sc}"></div>
   <div class="area-name" title="{a['area']}">{a['area']}</div>
   <div class="area-counts">{a['offline']}/{a['total']} off</div>
   <div class="mini-bar"><div class="mfill" style="width:{min(a['pct'],100)}%;background:{bc}"></div></div>
   <div class="area-pct {pc}">{pct_str}</div>
   {'<span class="cam-arrow" id="arr_'+uid+'">▶</span>' if show_cameras and a.get('cameras') else ''}
  </div>
  {cam_panel}
 </div>"""

CSS = """
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;background:#0f1117;color:#e0e0e0;min-height:100vh}
.top-bar{background:linear-gradient(135deg,#1a1f2e,#232a3e);padding:18px 32px;
 display:flex;align-items:center;justify-content:space-between;border-bottom:2px solid #2d3550}
.logo{font-size:2rem;font-weight:900;letter-spacing:3px}
.sublabel{font-size:0.88rem;letter-spacing:1px;font-weight:600;margin-top:2px}
.fecha{font-size:0.82rem;color:#8899bb;margin-top:3px}
.top-bar h1{font-size:1.2rem;font-weight:700;color:#fff}
.manual-badge{display:inline-block;background:#f59e0b22;color:#f59e0b;
 border:1px solid #f59e0b55;border-radius:6px;padding:3px 10px;
 font-size:0.75rem;font-weight:700;margin-left:10px;vertical-align:middle}
.tab-bar{display:flex;gap:4px;padding:10px 32px 0;background:#13172a;
 border-bottom:2px solid #2d3550;flex-wrap:wrap}
.tab-btn{padding:8px 16px;border:none;border-radius:8px 8px 0 0;cursor:pointer;
 font-size:0.78rem;font-weight:600;letter-spacing:0.5px;background:#0d1a30;
 color:#7a9cc0;border:1px solid #2d3550;border-bottom:none;transition:all .15s}
.tab-btn.active{background:#1a1f2e;color:#fff;border-color:#3a4a6e}
.tab-btn:hover:not(.active){background:#1a1f2e;color:#cde}
.tab-btn.manual{color:#f59e0b}
.tab-pane{display:none}.tab-pane.active{display:block}
.summary-bar{display:flex;gap:16px;padding:20px 32px;background:#13172a;
 border-bottom:1px solid #2d3550;flex-wrap:wrap}
.s-card{flex:1;min-width:130px;background:#1a1f2e;border-radius:10px;
 padding:14px 18px;border-left:4px solid}
.s-card .lbl{font-size:0.72rem;color:#8899bb;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px}
.s-card .val{font-size:1.7rem;font-weight:800}
.s-card .sub{font-size:0.78rem;color:#aab;margin-top:2px}
.sem-summary{display:flex;gap:12px;padding:14px 32px 4px;flex-wrap:wrap}
.sem-badge{display:flex;align-items:center;gap:8px;background:#1a1f2e;
 border-radius:8px;padding:9px 16px;font-size:0.83rem;font-weight:600}
.dot{width:13px;height:13px;border-radius:50%;flex-shrink:0}
.dot-red{background:#ef4444;box-shadow:0 0 8px #ef444499}
.dot-yellow{background:#f59e0b;box-shadow:0 0 8px #f59e0b99}
.dot-green{background:#22c55e;box-shadow:0 0 8px #22c55e99}
.clients-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(460px,1fr));
 gap:22px;padding:22px 32px 32px}
.client-panel{background:#1a1f2e;border-radius:14px;overflow:hidden;border:1px solid #2d3550}
.client-header{padding:16px 22px;display:flex;align-items:center;justify-content:space-between}
.client-header h2{font-size:1.05rem;font-weight:700;color:#fff}
.client-stats{display:flex;gap:10px}
.stat-b{display:flex;flex-direction:column;align-items:center;
 background:rgba(0,0,0,0.3);border-radius:8px;padding:6px 14px}
.stat-b .n{font-size:1.25rem;font-weight:800}
.stat-b .l{font-size:0.65rem;color:#99aacc;text-transform:uppercase}
.pct-bar-wrap{margin:0 22px 14px;background:#0f1117;border-radius:6px;overflow:hidden;height:8px}
.pct-bar-fill{height:100%;border-radius:6px;background:linear-gradient(90deg,#ef4444,#f59e0b)}
.legend{display:flex;gap:16px;padding:4px 22px 12px;font-size:0.72rem;color:#8899bb}
.legend-item{display:flex;align-items:center;gap:5px}
.areas-list{padding:0 16px 16px;display:flex;flex-direction:column;gap:6px;
 max-height:540px;overflow-y:auto}
.areas-list::-webkit-scrollbar{width:5px}
.areas-list::-webkit-scrollbar-track{background:#0f1117}
.areas-list::-webkit-scrollbar-thumb{background:#334;border-radius:3px}
.area-row{display:flex;align-items:center;gap:10px;background:#131826;
 border-radius:8px;padding:8px 12px;border:1px solid #222b40}
.semaforo{width:13px;height:13px;border-radius:50%;flex-shrink:0}
.sem-green{background:#22c55e;box-shadow:0 0 7px #22c55e88}
.sem-yellow{background:#f59e0b;box-shadow:0 0 7px #f59e0b88}
.sem-red{background:#ef4444;box-shadow:0 0 7px #ef444488}
.area-name{flex:1;font-size:0.8rem;color:#ccd;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.area-counts{font-size:0.75rem;color:#8899bb;white-space:nowrap}
.area-pct{font-size:0.8rem;font-weight:700;width:44px;text-align:right}
.pct-red{color:#ef4444}.pct-yellow{color:#f59e0b}.pct-green{color:#22c55e}
.mini-bar{width:65px;background:#0f1117;border-radius:3px;height:5px;flex-shrink:0}
.mini-bar .mfill{height:100%;border-radius:3px}
.hist-section{padding:18px 32px 8px}
.hist-title{font-size:0.78rem;font-weight:700;color:#556;text-transform:uppercase;
 letter-spacing:1px;margin-bottom:12px}
.hist-row{display:flex;gap:6px;align-items:flex-end;margin-bottom:14px}
.hist-label{font-size:0.72rem;color:#7a9cc0;width:90px;flex-shrink:0;padding-bottom:18px}
.hist-bars{display:flex;gap:5px;align-items:flex-end;flex:1}
.hist-bar-wrap{display:flex;flex-direction:column;align-items:center;gap:3px;flex:1}
.hist-bar-bg{width:100%;background:#131826;border-radius:3px 3px 0 0;
 position:relative;overflow:hidden;height:50px;border:1px solid #1a2e50}
.hist-bar-fill{position:absolute;bottom:0;width:100%;border-radius:2px 2px 0 0}
.hist-date{font-size:0.62rem;color:#445;text-align:center;white-space:nowrap}
.hist-pct{font-size:0.68rem;font-weight:700;text-align:center}
.hist-bar-bg.manual-bar{border-color:#f59e0b55}
.hist-date.manual-date{color:#f59e0b99}
.footer{text-align:center;padding:14px;color:#445;font-size:0.75rem}
.area-wrap{display:flex;flex-direction:column}
.cam-arrow{font-size:0.7rem;color:#556;margin-left:4px;transition:transform .2s;flex-shrink:0}
.cam-arrow.open{transform:rotate(90deg);color:#4a9eff}
.cam-panel{display:none;padding:6px 12px 8px 36px;background:#0d1220;
 border-left:2px solid #1a2e50;margin:0 0 4px 0;border-radius:0 0 6px 6px}
.cam-item{display:flex;align-items:center;gap:7px;padding:3px 0;border-bottom:1px solid #131826}
.cam-item:last-child{border-bottom:none}
.cam-dot{width:9px;height:9px;border-radius:50%;flex-shrink:0}
.cam-on{background:#22c55e;box-shadow:0 0 5px #22c55e88}
.cam-off{background:#ef4444;box-shadow:0 0 5px #ef444488}
.cam-name{font-size:0.75rem;color:#aab;line-height:1.5}
.cam-name-off{color:#ef4444}
.cam-addr{font-size:0.68rem;color:#556;font-family:monospace;margin-left:2px}
@media(max-width:900px){.clients-grid{grid-template-columns:1fr}.tab-bar{padding:10px 16px 0}}
"""

TAB_JS = """<script>
function showTab(id,btn){
  document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
}
function toggleCams(uid){
  var p=document.getElementById('cams_'+uid);
  var a=document.getElementById('arr_'+uid);
  if(!p)return;
  if(p.style.display==='block'){p.style.display='none';if(a)a.classList.remove('open');}
  else{p.style.display='block';if(a)a.classList.add('open');}
}
</script>"""

# ─────────────────────────────────────────────
# build_panel
# ─────────────────────────────────────────────
def build_panel(label, hdr_color, subset, show_cameras=False):
    total   = sum(a['total']   for a in subset)
    offline = sum(a['offline'] for a in subset)
    online  = total - offline
    pct     = round(offline / total * 100, 1) if total else 0
    rows    = '\n'.join(area_row(a, show_cameras)
                        for a in sorted(subset, key=lambda x: -x['pct']))
    return f"""
 <div class="client-panel">
  <div class="client-header" style="background:linear-gradient(135deg,{hdr_color},#1a1f2e)">
   <div>
    <h2>{label}</h2>
    <div style="font-size:0.78rem;color:#8899bb;margin-top:3px">{len(subset)} áreas · {total} cámaras</div>
   </div>
   <div class="client-stats">
    <div class="stat-b"><span class="n" style="color:#22c55e">{online}</span><span class="l">Online</span></div>
    <div class="stat-b"><span class="n" style="color:#ef4444">{offline}</span><span class="l">Offline</span></div>
    <div class="stat-b"><span class="n" style="color:#f59e0b">{pct}%</span><span class="l">Caída</span></div>
   </div>
  </div>
  <div class="pct-bar-wrap"><div class="pct-bar-fill" style="width:{min(pct,100)}%"></div></div>
  <div class="legend">
   <div class="legend-item"><div class="semaforo sem-green"></div> = 0%</div>
   <div class="legend-item"><div class="semaforo sem-yellow"></div> &lt;20%</div>
   <div class="legend-item"><div class="semaforo sem-red"></div> ≥20%</div>
  </div>
  <div class="areas-list">{rows}</div>
 </div>"""

# ─────────────────────────────────────────────
# Sparkbars
# ─────────────────────────────────────────────
def build_hist_section(history, panels_config):
    if len(history) < 2: return ""
    rows_html = []
    for client_keys, label, color in panels_config:
        bars = []
        for fecha_dt, areas, is_manual in history:
            subset = [a for a in areas if a['client'] in client_keys]
            if not subset: continue
            total   = sum(a['total']   for a in subset)
            offline = sum(a['offline'] for a in subset)
            pct     = round(offline / total * 100, 1) if total else 0
            _, _, bar_color = sem(pct)
            height  = min(pct * 2, 100)
            day_lbl = fecha_dt.strftime("%d/%m").lstrip("0").replace("/0", "/")
            manual_class = " manual-bar"  if is_manual else ""
            date_class   = " manual-date" if is_manual else ""
            cam_icon     = " 📷"           if is_manual else ""
            bars.append(f"""
  <div class="hist-bar-wrap">
   <div class="hist-bar-bg{manual_class}">
    <div class="hist-bar-fill" style="height:{height}%;background:{bar_color}"></div>
   </div>
   <div class="hist-pct" style="color:{bar_color}">{pct}%</div>
   <div class="hist-date{date_class}">{day_lbl}{cam_icon}</div>
  </div>""")
        short = label.split('—')[-1].strip() if '—' in label else label
        rows_html.append(f"""
 <div class="hist-row">
  <div class="hist-label" style="color:{color}">{short}</div>
  <div class="hist-bars">{''.join(bars)}</div>
 </div>""")
    if not rows_html: return ""
    return f"""<div class="hist-section">
 <div class="hist-title">📈 Historial últimos {len(history)} días</div>
 {''.join(rows_html)}
</div>"""

# ─────────────────────────────────────────────
# build_dashboard  ← MODIFICADO: inyecta auth
# ─────────────────────────────────────────────
def build_dashboard(titulo, logo_txt, logo_color, sub_label, panels_config, history, filename):

    # Obtener config de auth para este archivo
    fname_key = os.path.basename(filename)
    auth_cfg  = DASHBOARD_AUTH.get(fname_key)

    def content_for_day(areas, fecha_str, is_today, history_for_hist, is_manual):
        all_subset = [a for a in areas if a['client'] in [c for cfg in panels_config for c in cfg[0]]]
        g_total   = sum(a['total']   for a in all_subset)
        g_offline = sum(a['offline'] for a in all_subset)
        g_online  = g_total - g_offline
        g_pct     = round(g_offline / g_total * 100, 1) if g_total else 0
        n_red     = sum(1 for a in all_subset if a['pct'] >= 20)
        n_yellow  = sum(1 for a in all_subset if 0 < a['pct'] < 20)
        n_green   = sum(1 for a in all_subset if a['pct'] == 0)
        manual_banner = ''
        if is_manual:
            manual_banner = (f'<div style="background:#f59e0b22;border:1px solid #f59e0b55;'
                             f'border-radius:8px;padding:10px 32px;margin:10px 32px 0;'
                             f'font-size:0.82rem;color:#f59e0b">📷 Datos de reporte manual '
                             f'exportado el {fecha_str}</div>')
        extra_cards = ''
        for client_keys, label, color in panels_config:
            sub = [a for a in areas if a['client'] in client_keys]
            if not sub: continue
            t = sum(a['total']   for a in sub)
            o = sum(a['offline'] for a in sub)
            p = round(o/t*100, 1) if t else 0
            short = label.split('—')[-1].strip() if '—' in label else label
            extra_cards += (f' <div class="s-card" style="border-color:{color};margin-left:6px">'
                            f'<div class="lbl">{short}</div>'
                            f'<div class="val" style="color:{color}">{t}</div>'
                            f'<div class="sub">{o} offline ({p}%)</div></div>\n')
        panels_html = '\n'.join(
            build_panel(lbl, color, [a for a in areas if a['client'] in cl_keys],
                        show_cameras=is_today)
            for cl_keys, lbl, color in panels_config
            if any(a['client'] in cl_keys for a in areas)
        )
        hist_html = build_hist_section(history_for_hist, panels_config) if is_today else ""
        return f"""{manual_banner}
<div class="summary-bar">
 <div class="s-card" style="border-color:{logo_color}"><div class="lbl">Total Cámaras</div>
  <div class="val" style="color:{logo_color}">{g_total}</div>
  <div class="sub">{len(all_subset)} áreas</div></div>
 <div class="s-card" style="border-color:#22c55e"><div class="lbl">Online</div>
  <div class="val" style="color:#22c55e">{g_online}</div>
  <div class="sub">{round(g_online/g_total*100,1) if g_total else 0}% disponibles</div></div>
 <div class="s-card" style="border-color:#ef4444"><div class="lbl">Offline</div>
  <div class="val" style="color:#ef4444">{g_offline}</div>
  <div class="sub">{g_pct}% fuera de línea</div></div>
 {extra_cards}
</div>
<div class="sem-summary">
 <div class="sem-badge"><div class="dot dot-red"></div><span style="color:#ef4444">{n_red}</span>&nbsp;áreas críticas (≥20%)</div>
 <div class="sem-badge"><div class="dot dot-yellow"></div><span style="color:#f59e0b">{n_yellow}</span>&nbsp;áreas en alerta (&lt;20%)</div>
 <div class="sem-badge"><div class="dot dot-green"></div><span style="color:#22c55e">{n_green}</span>&nbsp;áreas sin caída (0%)</div>
</div>
{hist_html}
<div class="clients-grid">{panels_html}</div>"""

    prefix     = re.sub(r'[^a-z]', '', titulo.lower())[:4]
    tabs_bar   = []
    tabs_content = []

    for i, (fecha_dt, areas, is_manual) in enumerate(reversed(history)):
        is_today = (i == 0)
        tab_id   = f"{prefix}_d{i}"
        day_lbl  = fecha_dt.strftime("%d/%m").lstrip("0").replace("/0", "/")
        cam_icon = " 📷" if is_manual else ""
        btn_label = f"Hoy {day_lbl}{cam_icon}" if is_today else f"{day_lbl}{cam_icon}"
        active    = "active" if is_today else ""
        manual_cls = " manual" if is_manual else ""
        fecha_str  = fecha_dt.strftime("%d/%m/%Y")
        tabs_bar.append(
            f'<button class="tab-btn {active}{manual_cls}" onclick="showTab(\'{tab_id}\',this)">{btn_label}</button>'
        )
        tabs_content.append(
            f'<div id="{tab_id}" class="tab-pane {active}">' +
            content_for_day(areas, fecha_str, is_today, history, is_manual) +
            '</div>'
        )

    today_fecha  = history[-1][0].strftime("%d/%m/%Y")
    today_manual = history[-1][2]
    manual_header = ' <span class="manual-badge">📷 Reporte Manual</span>' if today_manual else ''

    # CSS: base + auth overlay
    full_css = CSS + (AUTH_CSS if auth_cfg else "")

    # Bloque de auth (se inyecta justo al abrir <body>)
    auth_html = ""
    if auth_cfg:
        hash_val, session_key, a_logo, a_color, a_sub = auth_cfg
        auth_html = auth_block(hash_val, session_key, a_logo, a_color, a_sub)

    html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{titulo} — Dashboard Cámaras</title>
<style>{full_css}</style></head><body>
{auth_html}
<div class="top-bar">
 <div>
  <div class="logo" style="color:{logo_color}">{logo_txt}{manual_header}</div>
  <div class="sublabel" style="color:{logo_color}88">{sub_label}</div>
  <div class="fecha">Dashboard de Estado de Cámaras — {today_fecha}</div>
 </div>
 <h1>Monitor de Disponibilidad</h1>
</div>
<div class="tab-bar">{''.join(tabs_bar)}</div>
{''.join(tabs_content)}
<div class="footer">Datos al {today_fecha} · Reporte automático Hik-Central Professional</div>
{TAB_JS}
</body></html>"""

    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f" ✓ {os.path.basename(filename)}{' (🔒 protegido)' if auth_cfg else ''}")

# ─────────────────────────────────────────────
# GIT
# ─────────────────────────────────────────────
def subir_github(repo_dir, fecha):
    try:
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        subprocess.run(['git', 'add', '.'], check=True)
        result = subprocess.run(['git', 'commit', '-m', f'Dashboard {fecha}'],
                                capture_output=True, text=True)
        if 'nothing to commit' in result.stdout or 'nothing to commit' in result.stderr:
            print(" ✓ GitHub ya estaba actualizado (sin cambios nuevos)")
            return
        subprocess.run(['git', 'pull', '--rebase', 'origin', 'main'], check=True)
        subprocess.run(['git', 'push'], check=True)
        print(" ✓ Subido a GitHub Pages correctamente")
    except subprocess.CalledProcessError as e:
        print(f" ERROR al subir a GitHub: {e}")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 55)
    print(" GENERADOR DE DASHBOARDS — KAIBIL / SEGURA v2.2")
    print("=" * 55)
    print("\n[1/4] Buscando Excels (hasta 7 días)...")
    pares = encontrar_excels(CARPETA_EXCEL)
    print(f" {len(pares)} archivo(s) encontrado(s):")
    for dt, ruta in pares:
        tipo = "📷 manual" if es_reporte_camara(ruta) else "📋 log"
        print(f"  {dt.strftime('%d/%m/%Y')} {tipo}  {os.path.basename(ruta)}")

    print("\n[2/4] Procesando datos...")
    history = []
    for fecha_dt, ruta in pares:
        areas, _ = leer_excel(ruta)
        is_manual = es_reporte_camara(ruta)
        history.append((fecha_dt, areas, is_manual))
        tipo = "manual" if is_manual else "log"
        print(f" ✓ {fecha_dt.strftime('%d/%m/%Y')} [{tipo}] — {len(areas)} áreas")

    today_fecha = history[-1][0].strftime("%d/%m/%Y")

    print("\n[3/4] Generando dashboards...")
    out = os.path.dirname(os.path.abspath(__file__))

    build_dashboard('Kaibil', 'KAIBIL', '#4a9eff', 'Seguridad Privada',
        [(['KS'], '🔵 KS — Kaibil Abonados',  '#1a3a6e'),
         (['KC'], '🟢 KC — Kaibil Barriales', '#1a4a2e')],
        history, os.path.join(out, 'index.html'))

    build_dashboard('Segura', 'SEGURA', '#a855f7', 'Monitoreo y Vigilancia',
        [(['SG'],         '🟣 SG — Segura',               '#3b1a6e'),
         (['SR','MK','MS'],'🩷 SR — Segura + Otros',       '#6e1a3a'),
         (['SP'],         '🔵 SP — Segura Punta del Este', '#1a3a7e')],
        history, os.path.join(out, 'segura.html'))

    build_dashboard('Segura Punta del Este', 'SEGURA', '#4a9eff', 'Punta del Este',
        [(['SP'], '🔵 SP — Segura Punta del Este', '#1a3a7e')],
        history, os.path.join(out, 'punta-del-este.html'))

    print("\n[4/4] Subiendo a GitHub...")
    subir_github(CARPETA_REPO, today_fecha)

    print("\n" + "=" * 55)
    print(" ✓ PROCESO COMPLETADO")
    print(f"\n 🔵 Kaibil:  https://{GITHUB_USUARIO}.github.io/dashboards/")
    print(f" 🟣 Segura:  https://{GITHUB_USUARIO}.github.io/dashboards/segura.html")
    print(f" 🔵 PDE:     https://{GITHUB_USUARIO}.github.io/dashboards/punta-del-este.html")
    print("=" * 55)

if __name__ == '__main__':
    main()
